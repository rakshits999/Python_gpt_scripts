import openpyxl
import openai

openai.api_key = "ENTER_YOUR_OPENAI_API_KEY"

workbook = openpyxl.load_workbook('excel.xlsx')
topics_sheet = workbook['Topics']
instructions_sheet = workbook['ChatGPT Instructions']

for i, row in enumerate(topics_sheet.iter_rows(min_row=2, max_col=2, values_only=True)):
    # if i <=3:
        topic_text = row[1]
        print(topic_text)
        column_number = 3  
        for r in instructions_sheet.iter_rows(min_row=2, max_col=4, max_row=2,  values_only=True):
            if "INSERT TOPIC" in r[3]:
                Linkedin = r[3].replace("INSERT TOPIC", topic_text)
                print(Linkedin)
                response = openai.Completion.create(
                    engine="text-davinci-003",
                    prompt=Linkedin,
                    max_tokens=2000,
                    n=1,
                    stop=None,
                    temperature=0.5,
                )
                print(response.choices[0].text.strip())
                topics_sheet.cell(row=i+2, column=column_number).value = response.choices[0].text.strip()
                column_number += 1  
        workbook.save('excel.xlsx')
                
        column_number = 4 
        for r in instructions_sheet.iter_rows(min_row=3, max_col=4, max_row=3, values_only=True):
            if "INSERT TOPIC" in r[3]:
                Youtube_Video = r[3].replace("INSERT TOPIC", topic_text)
                print(Youtube_Video)
                response = openai.Completion.create(
                    engine="text-davinci-003",
                    prompt=Youtube_Video,
                    max_tokens=2000,
                    n=1,
                    stop=None,
                    temperature=0.5,
                )
                print(response.choices[0].text.strip())
                topics_sheet.cell(row=i+2, column=column_number).value = response.choices[0].text.strip()
                column_number += 1  
        workbook.save('excel.xlsx')
                
        column_number = 5  
        for r in instructions_sheet.iter_rows(min_row=4, max_col=4, max_row=4, values_only=True):
            if "INSERT TOPIC" in r[3]:
                Youtube_Short = r[3].replace("INSERT TOPIC", topic_text)
                print(Youtube_Short)
                response = openai.Completion.create(
                    engine="text-davinci-003",
                    prompt=Youtube_Short,
                    max_tokens=2000,
                    n=1,
                    stop=None,
                    temperature=0.5,
                )
                print(response.choices[0].text.strip())
                topics_sheet.cell(row=i+2, column=column_number).value = response.choices[0].text.strip()
                column_number += 1  

        workbook.save('excel.xlsx')
                
        column_number = 6 
        for r in instructions_sheet.iter_rows(min_row=5, max_col=4, max_row=5, values_only=True):
            if "INSERT TOPIC" in r[3]:
                Email_Youtube = r[3].replace("INSERT TOPIC", topic_text)
                print(Email_Youtube)
                response = openai.Completion.create(
                    engine="text-davinci-003",
                    prompt=Email_Youtube,
                    max_tokens=2000,
                    n=1,
                    stop=None,
                    temperature=0.5,
                )
                print(response.choices[0].text.strip())
                topics_sheet.cell(row=i+2, column=column_number).value = response.choices[0].text.strip()
                column_number += 1
        
        workbook.save('excel.xlsx')
        
        column_number = 7
        for r in instructions_sheet.iter_rows(min_row=6, max_col=4, max_row=6, values_only=True):
            if "INSERT TOPIC" in r[3]:
                Email_LinkedIN = r[3].replace("INSERT TOPIC", topic_text)
                print(Email_LinkedIN)
                response = openai.Completion.create(
                    engine="text-davinci-003",
                    prompt=Email_LinkedIN,
                    max_tokens=2000,
                    n=1,
                    stop=None,
                    temperature=0.5,
                )
                print(response.choices[0].text.strip())
                topics_sheet.cell(row=i+2, column=column_number).value = response.choices[0].text.strip()
                column_number += 1
        
        workbook.save('excel.xlsx')
